# -*- coding: utf-8 -*-
"""
Created on Thu Oct  5 11:56:49 2023

@author: 1148055
"""
try:
    from backend_modules.xlFuncs import xlFuncs
except:
    from xlFuncs import xlFuncs
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Border, Side

class xlIntrfc(xlFuncs):
    def __init__(self,fpath='',frmt=True):
        self.fpath = fpath
        self.frmt=frmt
    
    def addIdx(self,ws,df=pd.DataFrame(), index=False, use_cols=True, col=1,row=1,border=False):
        """copy a range onto a page and format the page"""
        
        if isinstance(df, pd.Series):
            df = pd.DataFrame(df)
        
        # remove the references to the ws the df is in - for traceability
        df = self.purgeWsRef(ws,df)
        
        # assign df value to the sheet
        self.dfToXl(ws, df, col, row, index, use_cols)
        
        # check for values with @ symbols
        df_form = self.getWsDataOpxl(ws)
        if isinstance(df_form,list) or isinstance(df_form,tuple):
            df_form = pd.DataFrame(df_form)
        
        return 
    
    def insrIdx(self,ws,map_dic,border=False):
        """inserts cells into an already existing worksheet
        map_dic={<address>:<value>}"""
        
        border_min_row = -1
        border_min_col = -1
        border_max_row = -1
        border_max_col = -1
        
        for addr in map_dic.keys():
            val = map_dic[addr]
            # robust way
            row_idx, col_idx = self.addrToIdcs(addr)
            
            if border_min_row == -1:
                border_min_row = row_idx
            else:
                border_min_row = min(row_idx,border_min_row)
            if border_max_row == -1:
                border_max_row = row_idx
            else:
                border_max_row = max(row_idx,border_max_row)
            if border_min_col == -1:
                border_min_col = col_idx
            else:
                border_min_col = min(col_idx,border_min_col)
            if border_max_col == -1:
                border_max_col = col_idx
            else:
                border_max_col = max(col_idx,border_max_col)
            
            ws.cell(row=int(row_idx), column=col_idx).value = val
        
        if border:
            
            self.setBorder(ws, border_min_row,border_min_col,border_max_row,border_max_col)
        
        return
    
    def dfToXl(self, ws, df, start_col, start_row, use_index, use_cols):
        """Adds a df to an excel worksheet where the top left corner is 
        specified by col_num and row_num"""
        
        temp = df.copy()
        if use_index:
            temp = temp.reset_index()
            # get rid of the index title if the index isn't named
            idx_name = 'index'
            if idx_name in temp.columns:
                temp = temp.rename(columns={idx_name:""})
    
        rows = list(dataframe_to_rows(temp,index=False,header=use_cols))
        
        
        # Write the DataFrame to the specified range
        for r_idx, row in enumerate(rows, start_row):
            for c_idx, value in enumerate(row, start_col):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
        
        return

    
    def setACells(self,ws,df,a_cells):
        """sets all of the @ cells correctly"""
        
        if len(a_cells) > 0:
            df_ref = self.wsToCellRef(ws,df,offs=-1) # ws header is the cols
            for cell in a_cells:
                ref = df_ref.iloc[cell]
                val = df.iloc[cell].replace("@","")
                # use formula 2 when dealing with arrays calcs that trip
                ws.range(ref).formula2 = val
        
        return
    
    def getACells(self,ws,df):
        """Returns a list of indices to the df and df_ref with the @ symbol in a formula"""
        
        a_cells = []
        
        def appendACells(srs, a_cells):
            a = srs.str.contains("=", na=False) & srs.str.contains("@", na=False)
            for idx in list(np.flatnonzero(a)):
                a_cells.append((idx,srs.name))
                
            return
        
        # determine which axis to apply to
        if len(df) < len(df.columns):
            # vertical axis
            df.apply(lambda row: appendACells(row, a_cells), axis = 1) 
        else:
            # horizontal axis
            df.apply(lambda col: appendACells(col, a_cells), axis = 0) 
        
        return a_cells
        
    def frmtWs(self,ws,dllr_cells=[],prc_cells=[],date_cells=[],bold_cells=[],ctr_cells=[],highl_cells=[],space_cells=[],dec_cells=[],data_dfs=[],str_cells=[],arr_cells=[],mrg_rngs=[],comm_cells=[],fontcolor_dic={},fontsize_dic={},perc_len=0,dec_len=2,dllr_prc=False,revise=False,border=False):
        """format a ws after data has been added"""
        
        if self.frmt:

            # verify formatting is correct
            if not isinstance(dllr_cells, list):
                dllr_cells = [dllr_cells]
            if not isinstance(prc_cells, list):
                prc_cells = [prc_cells]
            if not isinstance(date_cells, list):
                date_cells = [date_cells]
            if not isinstance(bold_cells, list):
                bold_cells = [bold_cells]        
            if not isinstance(ctr_cells, list):
                ctr_cells = [ctr_cells]    
            if not isinstance(highl_cells, list):
                highl_cells = [highl_cells]    
            if not isinstance(space_cells, list):
                space_cells = [space_cells]    
            if not isinstance(dec_cells, list):
                dec_cells = [dec_cells]    
            if not isinstance(data_dfs, list):
                data_dfs = [data_dfs]    
            if not isinstance(str_cells, list):
                str_cells = [str_cells]    
            if not isinstance(arr_cells, list):
                arr_cells = [arr_cells]    
            if not isinstance(mrg_rngs, list):
                mrg_rngs = [mrg_rngs]    
            if not isinstance(comm_cells, list):
                comm_cells = [comm_cells]    
            
            
            if not revise:
                # set everything to a number with no decimal places
                all_rng = self.opxlUsedRange(ws)
                self.frmtGenr(ws, all_rng)
            
            # format dollars
            for cell in dllr_cells:
                self.frmtDlr(ws, cell, dllr_prc)
            
            # format percent
            for cell in prc_cells:
                self.frmtPrc(ws, cell, perc_len)
            
            # format dates
            for cell in date_cells:
                self.frmtDate(ws, cell)
            
            # format bold
            for cell in bold_cells:
                self.frmtBold(ws, cell)
            
            for cell in highl_cells:
                self.frmtHighlight(ws,cell)
            
            for cell in space_cells:
                self.frmtSpace(ws,cell)
            
            # imbue with color for controller cells
            for cell in ctr_cells:
                self.frmtCntrllr(ws, cell)
            
            # format decimal
            for cell in dec_cells:
                self.frmtDec(ws, cell, dec_len)
             
            # format data
            self.dark = True
            for df in data_dfs:
                self.frmtData(ws, df,border=border)
             
            # format string
            for cell in str_cells:
                self.frmtStr(ws, cell)
                
            # merge cells
            for rng in mrg_rngs:
                self.frmtMrg(ws,rng)
            
            # comma cells
            for cell in comm_cells:
                self.frmtComma(ws, cell)
            
            # change fontcolor
            for cell_addr in fontcolor_dic.keys():
                fontcolor = fontcolor_dic[cell_addr]
                self.changeFontcolor(ws, cell_addr, fontcolor)
            
            # change fontsize
            for cell_addr in fontsize_dic.keys():
                fontsize = fontsize_dic[cell_addr]
                self.changeFontsize(ws, cell_addr, fontsize)
            
            # format array
            for cell in arr_cells:
                self.frmtArr(ws, cell)
            
        return
    
    def frmtDlr(self,ws,rng,dec=False):
        """modifies cell range to be in dollar format"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # format a number as $
        number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
        if dec:
            number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        # set the number format
        self.setNmbrFrmt(ws, rng, number_format)
        
        return
    
    def frmtPrc(self,ws,rng,perc_len=0):
        """modifies cell range to be in percent format"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # create a number format
        base = '0'
        if perc_len > 0:
            base = base + "."
            for i in range(perc_len):
                base = base + "0"
        number_format = f'{base}%'
        
        # set the number format
        self.setNmbrFrmt(ws, rng, number_format)
        
        return
        
    def frmtDate(self, ws, rng):
        """"modify cell range to date format mm/dd/yyyy"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # create anumber format
        number_format = 'M/D/YYYY'
        
        # set the number format
        self.setNmbrFrmt(ws, rng, number_format)
        
        return
    
    def frmtData(self,ws,df,border=False):
        """Gives a generic color to cells that should be highlighted not not edited"""
        
        # gray
        color = (217,217,217)
        
        # get the df range
        df_rng = self.dfToRange(ws,df)
        
        # set the color
        self.setRngColor(ws,df_rng,color,grid=True)
        
        # set the border
        if border:
            self.setAllBorders(ws,df_rng)
        
        return
    
    def setAllBorders(self, ws, rng):
        """Sets border for all cells in a range"""
        
        # remove the sheet header
        rng = rng.split("!")[-1]
        
        # iterate through all cells to add a border
        thin = Side(border_style="thin", color="000000")
        for row in ws[rng]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                
        return
    
    def frmtHighlight(self,ws,rng):
        """Gives a generic color to cells that should be highlighted not not edited"""
        
        # bluish gray
        color = (197,217,241)
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # set the color
        self.setRngColor(ws,rng,color)
        
        return
    
    def frmtSpace(self,ws,rng):
        """Gives a generic color to cells for spacing"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # bluish gray
        color = (238,236,225)
        
        # set the color
        self.setRngColor(ws,rng,color)
        
        return
    
    def frmtCntrllr(self,ws, rng):
        """modify a cell with a predefined color for controllers"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # create a color
        color = (252,228,214)
        
        # set the color
        self.setRngColor(ws,rng,color)
        
        return
    
    def frmtBold(self,ws,rng):
        """change the format to bold"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # Create a bold font
        bold_font = Font(bold=True)
        
        # set the bold font
        self.setFont(ws, rng, bold_font)
        
        return
    
    def frmtGenr(self,ws,rng):
        """Formats to no decimals and include commas"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # create a number format
        number_format = "##0"
        
        # set the number format
        self.setNmbrFrmt(ws, rng, number_format)
        
        return
    
    def frmtDec(self,ws,rng,dec_len):
        """Formats to no decimals and include commas"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # create a number format
        number_format = "0."
        for i in range(dec_len):
            number_format = number_format + '0'
        
        # set the number format
        self.setNmbrFrmt(ws, rng, number_format)
        
        return
    
    def frmtStr(self,ws,rng):
        """Formats to string"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # create a number format
        number_format = '@'
        
        # set the number format
        self.setNmbrFrmt(ws, rng, number_format)
        
        return
    
    def frmtComma(self, ws, rng):
        """"modify cell range to date format mm/dd/yyyy"""
        
        # verify the range is a string
        rng = self.verifRng(rng)
        
        # create anumber format
        number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        
        # set the number format
        self.setNmbrFrmt(ws, rng, number_format)
        
        return
    
    
    def frmtMrg(self,ws,rng):
        """Formats to string"""
        from openpyxl.styles import Alignment
        
        # split the range into start and stop cells
        rng = rng.split('!')[-1]
        start_cell = rng.split(':')[0]
        end_cell = rng.split(':')[1]
        
        # get the start col, start row, end col and end row
        start_col = self.getColIdx(start_cell)
        start_row = self.getRowIdx(start_cell)
        end_col = self.getColIdx(end_cell)
        end_row = self.getRowIdx(end_cell)
        
        # merge the range
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        
        # align center
        first_cell = ws.cell(start_row,start_col) #or currentCell = ws['A1']
        first_cell.alignment = Alignment(horizontal='center')
        
        return
    
    def changeFontsize(self,ws, rng, fontsize):
        """Change the original fontsize to something else, only accepts single cell range"""
        
        cell_addr = rng.split('!')[-1]
        row = self.getRowIdx(cell_addr)
        col = self.getColIdx(cell_addr)
        
        cell = ws.cell(row, col)
        cell.font = Font(size=fontsize)
        
        return
    
    def changeFontcolor(self,ws, rng, fontcolor):
        """Change the original fontcolor to something else, only accepts single cell range"""
        
        cell_addr = rng.split('!')[-1]
        row = self.getRowIdx(cell_addr)
        col = self.getColIdx(cell_addr)
        
        cell = ws.cell(row, col)
        color = self.rgbToHex(fontcolor)
        cell.font = Font(color=color)
        
        return
    
    def frmtArr(self,ws,rng):
        """Formats a range as an array"""
        
        start_cell, end_cell = self.rngToCells(rng)
        # Iterate over the cells in the range and set the fill color
        for row in ws[start_cell:end_cell]:
            for cell in row:
                addr = cell.coordinate
                curr_val = cell.value
                cell.value = ArrayFormula(addr, curr_val)
                # ws.formula_attributes[addr] = {'t': 'array', 'ref': f"{addr}:{addr}"}
            
        return
    
    def setRngColor(self, ws, rng, color, text=False, grid=False):
        """Sets the color to a specified range in openpyxl"""
        
        if isinstance(color,tuple):
            color = self.rgbToHex(color)
        
        # remove the sheet header
        # if len(rng.split("!")) > 1:
        rng = rng.split("!")[-1]
        
        # Split the cell range into start and end cells
        if ":" in rng:
            start_cell, end_cell = rng.split(':')
        else:
            start_cell = rng
            end_cell   = rng
            
        # Create a fill with the desired color
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        blank_fill = PatternFill(start_color= self.rgbToHex((255,255,255)), end_color=color, fill_type='solid') # white
        
        # alternate dark and light if it is a grid
        if grid:
            dark = False
            for row in ws[start_cell:end_cell]:
                for cell in row:
                    if dark:
                        if text:
                            cell.style.font.color.index = color
                        else:
                            cell.fill = fill
                    else:
                        if text:
                            cell.style.font.color.index = color
                        else:
                            cell.fill = blank_fill
                dark = not dark
            
        else:
            # Iterate over the cells in the range and set the fill color
            for row in ws[start_cell:end_cell]:
                for cell in row:
                    if text:
                        cell.style.font.color.index = color
                    else:
                        cell.fill = fill
                
        return
    
    def setNmbrFrmt(self, ws, rng, number_format):
        """Sets the number format in openpyxl"""
        
        start_cell, end_cell = self.rngToCells(rng)
            
        # Iterate over the cells in the range and set the fill color
        for row in ws[start_cell:end_cell]:
            for cell in row:
                cell.number_format = number_format
        
        return
    
    def setFont(self, ws, rng, font):
        """Sets the style parameter to a range in openpyxl"""
        
        start_cell, end_cell = self.rngToCells(rng)
        
        # Iterate over the cells in the range and set the fill color
        for row in ws[start_cell:end_cell]:
            for cell in row:
                cell.font = font
                
        return
    
    def rngToCells(self,rng):
        """Converts a range into the list of cells"""
        
        # remove the sheet header
        if len(rng.split("!")) > 1:
            rng = rng.split("!")[1]
        
        # Split the cell range into start and end cells
        if ":" in rng:
            start_cell, end_cell = rng.split(':')
        else:
            start_cell = rng
            end_cell   = rng
        
        return start_cell, end_cell
        
    
    def rgbToHex(self,rgb):
        """Converts an rgb tuple to hex"""
        
        hex_color = '#%02x%02x%02x' % rgb
        
        hex_color = hex_color.replace("#","00").upper()
        
        return hex_color
    
    # def getDfVals(self,ws_opxl):
    #     """Returns a df of values from the worksheet.
        
    #     ************************************
    #     DO NOT USE IN PRODUCTION
    #     ************************************
    #     """
        
    #     # close the parent workbook
    #     wb = ws_opxl.parent
    #     wb.close()
        
    #     #us xlwings to get the data
    #     with xw.App(visible=False) as app:
    #         # open the workbook
    #         wb_xw = xw.Book(self.fpath)
            
    #         # get the worksheet
    #         ws_xw = wb_xw.sheets(ws_opxl.title)
            
    #         # get the dataframe of values
    #         df = self.getWsDataXw(ws_xw)
            
    #         # close the workbook
    #         wb_xw.close()
        
    #     # reopen with openpyxl
    #     wb = load_workbook(self.fpath)
        
    #     return wb, df
    
    def verifRng(self,rng):
        """makes sure the rng is a string"""
        
        if not isinstance(rng,str):
            rng = self.cellsToRange(rng)
        
        return rng
    
    def saveOpxl(self,ws_opxl):
        """waves a workbook after adding it or formatting it"""
        
        # save the workbook
        wb_opxl = ws_opxl.parent
        wb_opxl.save(self.fpath)
        
        return
    
    def setBorder(self, ws, border_min_row,border_min_col,border_max_row,border_max_col):
        """Set the border of a cell range"""
        thin = Side(border_style="thick", color="000000")
        cell_rng = f'{self.numToCol(border_min_col)}{border_min_row}:{self.numToCol(border_max_col)}{border_max_row}'
        for row in ws[cell_rng]:
            for cell in row:
                # row_idx, col_idx = self.addrToIdcs(cell.address)
                left = False
                right = False
                top = False 
                bottom = False
                if cell.row == border_min_row:
                    top = True
                    
                if cell.row == border_max_row:
                    bottom = True
                    
                if cell.column == border_min_col:
                    left = True
                    
                if cell.column == border_max_col:
                    right = True
                
                if top and right:
                    cell.border = Border(top=thin,right=thin)
                elif top and left:
                    cell.border = Border(top=thin,left=thin)
                elif bottom and right:
                    cell.border = Border(bottom=thin,right=thin)
                elif bottom and left:
                    cell.border = Border(bottom=thin,left=thin)
                elif top:
                    cell.border = Border(top=thin)
                elif bottom:
                    cell.border = Border(bottom=thin)
                elif left:
                    cell.border = Border(left=thin)
                elif right:
                    cell.border = Border(right=thin)